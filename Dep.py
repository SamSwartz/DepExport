from urllib.parse import parse_qs, urlencode
import requests
import openpyxl
import pprint

# Step 1: Obtain API credentials
client_id = '4e97513dc1aee76830dd43ba617eaccf97f87bef' 
client_secret = 'bd5eee02684a8e2456b3e3d94ed0b5f914c8bf29'

# Config
redirect_uri = 'https://samswartz.github.io/DepAPI/'

# Build authorization URL
auth_url = 'https://once.deputy.com/my/oauth/login'
params = {
  'client_id': client_id,
  'redirect_uri': redirect_uri,
  'response_type': 'code',
  'scope': 'longlife_refresh_token' 
}

auth_url += '?' + urlencode(params)



# Redirect user to auth URL
print(f'Please authorize at: {auth_url}')
# User authenticates and gets redirected back

# Wait for user input
redirect_url = input('Paste full redirect URL:')

# Extract code  
code = parse_qs(redirect_url)['code'][0]

print(f'Got code: {code}')

# Step 2: Connect to Deputy API
# Generate access token
token_url = 'https://lakeshomes.na.deputy.com/oauth/access_token'
data = {'grant_type': 'authorization_code',
        'code': 'auth_code',
        'scope': 'longlife_refresh_token'}
response = requests.post(token_url, data=data)
try:
  access_token = response.json()['access_token']
except KeyError:
  print("Access token not in response")
  access_token = None

api_url = 'https://lakeshomes.na.deputy.com/api/v1'
# Make API request with access token 
headers = {'Authorization': 'Bearer ' + access_token}
response = requests.get(api_url + '/timesheets', headers=headers)

# Step 3: Fetch timesheet data
#Get start and end dates
start_date = input("Enter start date (YYYY-MM-DD): ")
end_date = input("Enter end date (YYYY-MM-DD): ")

# Get approved timesheets for date range
params = {
  'status': 'approved',
  'startDateTime': start_date, 
  'endDateTime': end_date
}

response = requests.get('https://lakeshomes.na.deputy.com/v1/Resource/Timesheet',
                        params=params)
timesheets = response.json()

# Step 4: # Parse and extract payroll data
timesheet_data = [] 

for sheet in timesheets:
  
  if sheet['PayRuleApproved']:
  
    name = sheet['user']['name']  
    hours = sheet['timeApproved'] / 3600
    pay_rule = sheet['payRuleId']
    cost = sheet['cost']
    
    data = {'name': name, 'hours': hours, 
            'pay_rule': pay_rule, 'cost': cost}
            
    timesheet_data.append(data)
            
# Step 5: Export to Excel
wb = openpyxl.Workbook()
sheet = wb.active 

# Add column headers
sheet['A1'] = 'Name'
sheet['B1'] = 'Code' 
sheet['C1'] = 'Entries'
sheet['D1'] = 'Hours'
sheet['E1'] = 'Area'
sheet['F1'] = 'Shift'

for i, row in enumerate(timesheet_data):
    
    sheet.cell(i+2, 1).value = row['name']
    sheet.cell(i+2, 2).value = row['code']
    sheet.cell(i+2, 3).value = row['entries'] 
    sheet.cell(i+2, 4).value = row['hours']
    sheet.cell(i+2, 5).value = row['area']
    
    for j, entry in enumerate(row['entries']):
        sheet.cell(i+2, 6+j).value = entry['shift']
        
wb.save('timesheets.xlsx')

# Step 6: Automate process
# Schedule code to run daily with cron, Windows task scheduler etc.
